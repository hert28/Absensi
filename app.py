# app.py — Entry point Flask, semua route
# Komentar dalam Bahasa Indonesia sesuai konvensi (context.md bagian 12)

from flask import (Flask, request, jsonify, render_template,
                   redirect, url_for, session, flash)
from flask_cors import CORS
from flask_socketio import SocketIO, emit
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

# Timezone WIB (UTC+8) — digunakan di semua fungsi waktu
WIB = ZoneInfo('Asia/Jakarta')

def now_wib():
    """Dapatkan waktu sekarang dalam timezone WIB."""
    return datetime.now(WIB)
import os
import base64
import threading
import numpy as np
import cv2
import database as db
from config import (FLASK_HOST, FLASK_PORT, FLASK_SECRET_KEY,
                    SNAPSHOT_PATH, TOLERANSI_MENIT, DATASET_PATH,
                    CONFIDENCE_THRESHOLD, ANTI_SPOOFING_THRESHOLD,
                    ESP32_ENABLED, ESP32_IP, ESP32_PORT, ESP32_TIMEOUT,
                    MODEL_PATH)

# ── Inisialisasi Flask + SocketIO ─────────────────────────────
app = Flask(__name__)
app.secret_key = FLASK_SECRET_KEY
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Status kamera global
camera_state = {'active': False}

# Tracker verifikasi konsekutif — wajah harus dikenali N kali berturut-turut
_consecutive_tracker = {'user_id': None, 'count': 0}


# ══════════════════════════════════════════════════════════════
# DECORATOR: login_required
# Semua route kecuali /login dan /register wajib pakai ini
# ══════════════════════════════════════════════════════════════
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'admin_id' not in session:
            flash('Silakan login terlebih dahulu.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════
# AUTH: Login, Register, Logout
# ══════════════════════════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Halaman login admin."""
    # Jika sudah login, langsung ke dashboard
    if 'admin_id' in session:
        return redirect(url_for('dashboard'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            error = 'Username dan password wajib diisi.'
        else:
            admin = db.get_admin_by_username(username)
            if admin and check_password_hash(admin['password_hash'], password):
                # Login berhasil — simpan ke session
                session['admin_id'] = admin['id']
                session['username'] = admin['username']
                flash('Login berhasil!', 'success')
                return redirect(url_for('dashboard'))
            else:
                error = 'Username atau password salah.'

    # Tampilkan link register hanya jika belum ada admin sama sekali
    show_register = (db.hitung_admin() == 0)

    return render_template('login.html', error=error, show_register=show_register)


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Halaman register admin pertama.
    Hanya bisa diakses jika tabel admin masih kosong (0 record).
    Jika sudah ada admin, redirect ke /login.
    """
    # Cek apakah sudah ada admin
    if db.hitung_admin() > 0:
        flash('Admin sudah terdaftar. Silakan login.', 'error')
        return redirect(url_for('login'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')

        # Validasi input
        if not username or not password:
            error = 'Username dan password wajib diisi.'
        elif len(password) < 8:
            error = 'Password minimal 8 karakter.'
        elif password != confirm:
            error = 'Konfirmasi password tidak cocok.'
        else:
            # Hash password dan simpan
            hashed = generate_password_hash(password)
            admin_id = db.tambah_admin(username, hashed)
            if admin_id:
                # Langsung login setelah register
                session['admin_id'] = admin_id
                session['username'] = username
                flash('Akun admin berhasil dibuat!', 'success')
                return redirect(url_for('dashboard'))
            else:
                error = 'Gagal membuat akun. Username mungkin sudah dipakai.'

    return render_template('register_admin.html', error=error)


@app.route('/logout')
def logout():
    """Logout admin — hapus session."""
    session.clear()
    flash('Anda telah logout.', 'success')
    return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════

@app.route('/')
@login_required
def dashboard():
    """Halaman utama dashboard."""
    statistik = db.get_statistik_dashboard()
    absensi = db.get_absensi_hari_ini()
    return render_template('dashboard.html',
                           active_page='dashboard',
                           statistik=statistik,
                           absensi_hari_ini=absensi,
                           conf_threshold=CONFIDENCE_THRESHOLD,
                           spoof_threshold=ANTI_SPOOFING_THRESHOLD)


# ══════════════════════════════════════════════════════════════
# MANAJEMEN KELAS
# ══════════════════════════════════════════════════════════════

@app.route('/kelas')
@login_required
def kelas_index():
    """Daftar semua kelas."""
    daftar = db.get_semua_kelas()
    # Tambahkan jumlah mahasiswa per kelas
    for k in daftar:
        k['jumlah_mhs'] = db.hitung_mahasiswa_per_kelas(k['id'])
    return render_template('kelas/index.html',
                           active_page='kelas', daftar_kelas=daftar)


@app.route('/kelas/tambah', methods=['GET', 'POST'])
@login_required
def kelas_tambah():
    """Tambah kelas baru."""
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_kelas', '').strip()
        angkatan = request.form.get('angkatan', '').strip()
        if not nama or not angkatan:
            error = 'Nama kelas dan angkatan wajib diisi.'
        else:
            hasil = db.tambah_kelas(nama, angkatan, session.get('admin_id'))
            if hasil:
                flash('Kelas berhasil ditambahkan!', 'success')
                return redirect(url_for('kelas_index'))
            error = 'Gagal menambahkan kelas.'
    return render_template('kelas/form.html',
                           active_page='kelas', kelas=None, error=error)


@app.route('/kelas/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def kelas_edit(id):
    """Edit kelas."""
    kelas = db.get_kelas_by_id(id)
    if not kelas:
        flash('Kelas tidak ditemukan.', 'error')
        return redirect(url_for('kelas_index'))
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_kelas', '').strip()
        angkatan = request.form.get('angkatan', '').strip()
        if not nama or not angkatan:
            error = 'Nama kelas dan angkatan wajib diisi.'
        elif db.update_kelas(id, nama, angkatan):
            flash('Kelas berhasil diperbarui!', 'success')
            return redirect(url_for('kelas_index'))
        else:
            error = 'Gagal memperbarui kelas.'
    return render_template('kelas/form.html',
                           active_page='kelas', kelas=kelas, error=error)


@app.route('/kelas/hapus/<int:id>', methods=['POST'])
@login_required
def kelas_hapus(id):
    """Hapus kelas (CASCADE ke MK dan jadwal)."""
    if db.hapus_kelas(id):
        flash('Kelas berhasil dihapus.', 'success')
    else:
        flash('Gagal menghapus kelas. Mungkin masih ada mahasiswa terkait.', 'error')
    return redirect(url_for('kelas_index'))


@app.route('/kelas/<int:id>/matakuliah')
@login_required
def kelas_detail(id):
    """Detail kelas + daftar matakuliah."""
    kelas = db.get_kelas_by_id(id)
    if not kelas:
        flash('Kelas tidak ditemukan.', 'error')
        return redirect(url_for('kelas_index'))
    matakuliah = db.get_matakuliah_by_kelas(id)
    jumlah_mhs = db.hitung_mahasiswa_per_kelas(id)
    return render_template('kelas/detail.html',
                           active_page='kelas', kelas=kelas,
                           matakuliah=matakuliah, jumlah_mhs=jumlah_mhs)


# ══════════════════════════════════════════════════════════════
# MANAJEMEN MATAKULIAH
# ══════════════════════════════════════════════════════════════

@app.route('/matakuliah/tambah', methods=['GET', 'POST'])
@login_required
def matakuliah_tambah():
    """Tambah matakuliah baru."""
    kelas_id_param = request.args.get('kelas_id', type=int)
    kelas_asal = db.get_kelas_by_id(kelas_id_param) if kelas_id_param else None
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_mk', '').strip()
        kode = request.form.get('kode_mk', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        sks  = request.form.get('sks', 2, type=int)
        if not nama or not kode or not kid:
            error = 'Semua field wajib diisi.'
        else:
            hasil = db.tambah_matakuliah(nama, kode, kid, sks)
            if hasil:
                flash('Mata kuliah berhasil ditambahkan!', 'success')
                return redirect(url_for('kelas_detail', id=kid))
            error = 'Gagal menambahkan. Kode MK mungkin sudah dipakai.'
    return render_template('matakuliah/form.html', active_page='kelas',
                           mk=None, kelas_asal=kelas_asal,
                           daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/matakuliah/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def matakuliah_edit(id):
    """Edit matakuliah."""
    mk = db.get_matakuliah_by_id(id)
    if not mk:
        flash('Matakuliah tidak ditemukan.', 'error')
        return redirect(url_for('kelas_index'))
    kelas_asal = db.get_kelas_by_id(mk['kelas_id'])
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_mk', '').strip()
        kode = request.form.get('kode_mk', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        sks  = request.form.get('sks', 2, type=int)
        if db.update_matakuliah(id, nama, kode, kid, sks):
            flash('Mata kuliah berhasil diperbarui!', 'success')
            return redirect(url_for('kelas_detail', id=kid))
        error = 'Gagal memperbarui. Kode MK mungkin sudah dipakai.'
    return render_template('matakuliah/form.html', active_page='kelas',
                           mk=mk, kelas_asal=kelas_asal,
                           daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/matakuliah/hapus/<int:id>', methods=['POST'])
@login_required
def matakuliah_hapus(id):
    """Hapus matakuliah (CASCADE ke jadwal)."""
    mk = db.get_matakuliah_by_id(id)
    kelas_id = mk['kelas_id'] if mk else None
    if db.hapus_matakuliah(id):
        flash('Mata kuliah berhasil dihapus.', 'success')
    else:
        flash('Gagal menghapus mata kuliah.', 'error')
    return redirect(url_for('kelas_detail', id=kelas_id) if kelas_id
                    else url_for('kelas_index'))


# ══════════════════════════════════════════════════════════════
# MANAJEMEN JADWAL
# ══════════════════════════════════════════════════════════════

@app.route('/jadwal')
@login_required
def jadwal_index():
    """Daftar semua jadwal."""
    return render_template('jadwal/index.html', active_page='jadwal',
                           daftar_jadwal=db.get_semua_jadwal())


@app.route('/jadwal/tambah', methods=['GET', 'POST'])
@login_required
def jadwal_tambah():
    """Tambah jadwal baru."""
    error = None
    if request.method == 'POST':
        mk_id       = request.form.get('matakuliah_id', type=int)
        hari         = request.form.get('hari', '').strip()
        jam_mulai    = request.form.get('jam_mulai', '').strip()
        jam_selesai  = request.form.get('jam_selesai', '').strip()
        if not mk_id or not hari or not jam_mulai or not jam_selesai:
            error = 'Semua field wajib diisi.'
        else:
            # batas_terlambat dihitung otomatis di database.py
            hasil = db.tambah_jadwal(mk_id, hari, jam_mulai, jam_selesai)
            if hasil:
                flash('Jadwal berhasil ditambahkan!', 'success')
                return redirect(url_for('jadwal_index'))
            error = 'Gagal menambahkan jadwal.'
    return render_template('jadwal/form.html', active_page='jadwal',
                           daftar_mk=db.get_semua_matakuliah(),
                           daftar_kelas=db.get_semua_kelas(),
                           error=error)


@app.route('/jadwal/hapus/<int:id>', methods=['POST'])
@login_required
def jadwal_hapus(id):
    """Hapus jadwal."""
    if db.hapus_jadwal(id):
        flash('Jadwal berhasil dihapus.', 'success')
    else:
        flash('Gagal menghapus jadwal.', 'error')
    return redirect(url_for('jadwal_index'))


# ══════════════════════════════════════════════════════════════
# MANAJEMEN MAHASISWA
# ══════════════════════════════════════════════════════════════

@app.route('/mahasiswa')
@login_required
def mahasiswa_index():
    """Daftar semua mahasiswa."""
    filter_kelas = request.args.get('kelas_id', type=int)
    if filter_kelas:
        daftar = db.get_users_by_kelas(filter_kelas)
    else:
        daftar = db.get_semua_user()
    # Hitung foto per mahasiswa untuk status biometrik
    for m in daftar:
        folder = os.path.join(DATASET_PATH, str(m['id']))
        if os.path.isdir(folder):
            m['foto_count'] = len([f for f in os.listdir(folder) if f.endswith('.jpg')])
        else:
            m['foto_count'] = 0
    total_mhs = len(db.get_semua_user())
    bio_aktif = sum(1 for m in db.get_semua_user()
                    if os.path.isdir(os.path.join(DATASET_PATH, str(m['id'])))
                    and len(os.listdir(os.path.join(DATASET_PATH, str(m['id'])))) > 0)
    return render_template('mahasiswa/index.html', active_page='mahasiswa',
                           daftar_mahasiswa=daftar,
                           daftar_kelas=db.get_semua_kelas(),
                           filter_kelas=filter_kelas,
                           total_mhs=total_mhs, bio_aktif=bio_aktif)


@app.route('/mahasiswa/register', methods=['GET', 'POST'])
@login_required
def mahasiswa_register():
    """Form registrasi mahasiswa baru (data + foto kamera).

    POST dari form biasa (tanpa foto) → simpan data ke DB, redirect ke daftar.
    POST dari AJAX (api_foto_upload) → sudah ditangani route terpisah.
    """
    error = None
    if request.method == 'POST':
        nama     = request.form.get('nama', '').strip()
        nim      = request.form.get('nim', '').strip()
        kelas_id = request.form.get('kelas_id', type=int)

        if not nama or not nim or not kelas_id:
            error = 'Semua field (Nama, NIM, Kelas) wajib diisi.'
        elif db.nim_sudah_ada(nim):
            error = f'NIM {nim} sudah terdaftar.'
        else:
            user_id = db.tambah_user(nama, nim, kelas_id)
            if user_id:
                flash(f'Mahasiswa {nama} berhasil didaftarkan! Lanjutkan pengambilan foto biometrik.', 'success')
                return redirect(url_for('mahasiswa_index'))
            else:
                error = 'Gagal menyimpan. NIM mungkin sudah dipakai.'

    return render_template('mahasiswa/register.html', active_page='mahasiswa',
                           daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/mahasiswa/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def mahasiswa_edit(id):
    """Edit data mahasiswa."""
    mhs = db.get_user_by_id(id)
    if not mhs:
        flash('Mahasiswa tidak ditemukan.', 'error')
        return redirect(url_for('mahasiswa_index'))
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama', '').strip()
        nim  = request.form.get('nim', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        if not nama or not nim or not kid:
            error = 'Semua field wajib diisi.'
        elif db.update_user(id, nama, nim, kid):
            flash('Data mahasiswa berhasil diperbarui!', 'success')
            return redirect(url_for('mahasiswa_index'))
        else:
            error = 'Gagal memperbarui. NIM mungkin sudah dipakai.'
    return render_template('mahasiswa/edit.html', active_page='mahasiswa',
                           mhs=mhs, daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/mahasiswa/hapus/<int:id>', methods=['POST'])
@login_required
def mahasiswa_hapus(id):
    """Hapus mahasiswa + folder dataset foto."""
    mhs = db.get_user_by_id(id)
    if db.hapus_user(id):
        # Hapus folder foto dataset jika ada
        if mhs:
            folder = os.path.join(DATASET_PATH, str(id))
            if os.path.isdir(folder):
                import shutil
                shutil.rmtree(folder, ignore_errors=True)
        flash('Mahasiswa berhasil dihapus.', 'success')
    else:
        flash('Gagal menghapus mahasiswa.', 'error')
    return redirect(url_for('mahasiswa_index'))


# ══════════════════════════════════════════════════════════════
# TAHAP 8: REKAP ABSENSI + EXPORT
# ══════════════════════════════════════════════════════════════

@app.route('/absensi/rekap')
@login_required
def absensi_rekap():
    """Rekap absensi dengan filter kelas, MK, dan rentang tanggal."""
    # Ambil parameter filter dari query string
    kelas_id     = request.args.get('kelas_id', type=int)
    matakuliah_id = request.args.get('matakuliah_id', type=int)
    filter_dari  = request.args.get('dari') or None
    filter_sampai = request.args.get('sampai') or None

    # Data untuk dropdown filter
    daftar_kelas = db.get_semua_kelas()
    daftar_mk    = db.get_semua_matakuliah()

    # Ambil rekap + ringkasan berdasarkan filter
    rekap    = db.get_rekap_absensi(kelas_id, filter_dari, filter_sampai, matakuliah_id)
    ringkasan = db.get_ringkasan_rekap(kelas_id, filter_dari, filter_sampai, matakuliah_id)

    # Bangun query string untuk link export (teruskan filter yang sama)
    _params = []
    if kelas_id:      _params.append(f'kelas_id={kelas_id}')
    if matakuliah_id: _params.append(f'matakuliah_id={matakuliah_id}')
    if filter_dari:   _params.append(f'dari={filter_dari}')
    if filter_sampai: _params.append(f'sampai={filter_sampai}')
    filter_query = ('?' + '&'.join(_params)) if _params else ''

    return render_template('absensi/rekap.html',
                           active_page='rekap',
                           rekap=rekap,
                           ringkasan=ringkasan,
                           daftar_kelas=daftar_kelas,
                           daftar_mk=daftar_mk,
                           filter_kelas=kelas_id,
                           filter_mk=matakuliah_id,
                           filter_dari=filter_dari,
                           filter_sampai=filter_sampai,
                           filter_query=filter_query)


@app.route('/absensi/export')
@login_required
def absensi_export():
    """Export rekap absensi ke CSV atau Excel (.xlsx).

    Query params: format=csv|xlsx, kelas_id, matakuliah_id, dari, sampai
    """
    import io
    fmt           = request.args.get('format', 'csv').lower()
    kelas_id      = request.args.get('kelas_id', type=int)
    matakuliah_id = request.args.get('matakuliah_id', type=int)
    filter_dari   = request.args.get('dari') or None
    filter_sampai = request.args.get('sampai') or None

    rekap = db.get_rekap_absensi(kelas_id, filter_dari, filter_sampai, matakuliah_id)

    # ── Header kolom ──
    headers = ['Nama', 'NIM', 'Kelas', 'Mata Kuliah', 'Kode MK',
               'Hari', 'Tanggal', 'Waktu Absen', 'Status', 'Keterangan']

    def _row(a):
        return [
            a.get('nama', ''), a.get('nim', ''), a.get('nama_kelas', ''),
            a.get('nama_mk', ''), a.get('kode_mk', ''), a.get('hari', ''),
            str(a.get('tanggal', '')), str(a.get('waktu_absen', '') or '-'),
            a.get('status', ''), a.get('alasan', '') or '-'
        ]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if fmt == 'xlsx':
        # ── Export Excel ──
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from flask import send_file

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Rekap Absensi'

            # Header baris 1: judul
            ws.merge_cells('A1:J1')
            ws['A1'] = 'REKAP ABSENSI — SISTEM ABSENSI'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Header baris 2: kolom
            header_fill = PatternFill('solid', fgColor='1B2024')
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font = Font(bold=True, color='8ED5FF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Data
            STATUS_COLOR = {
                'hadir': 'D1FAE5', 'terlambat': 'FEF3C7',
                'izin': 'DBEAFE', 'sakit': 'FEF3C7', 'alpha': 'FEE2E2'
            }
            for row_idx, a in enumerate(rekap, 3):
                row_data = _row(a)
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    status = a.get('status', '')
                    # Warnai status (kolom ke-9)
                    if col_idx == 9 and status in STATUS_COLOR:
                        cell.fill = PatternFill('solid', fgColor=STATUS_COLOR[status])

            # Lebar kolom otomatis
            col_widths = [25, 15, 12, 25, 12, 10, 12, 14, 12, 25]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                             download_name=f'rekap_absensi_{timestamp}.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except ImportError:
            flash('openpyxl belum terinstall. Gunakan export CSV.', 'error')
            return redirect(url_for('absensi_rekap'))

    else:
        # ── Export CSV ──
        import csv
        from flask import Response

        def generate_csv():
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            for a in rekap:
                writer.writerow(_row(a))
            return buf.getvalue()

        return Response(
            generate_csv(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=rekap_absensi_{timestamp}.csv'}
        )


# ══════════════════════════════════════════════════════════════
# TAHAP 9: LAPORAN KEHADIRAN
# ══════════════════════════════════════════════════════════════

@app.route('/laporan')
@login_required
def laporan_index():
    """Laporan kehadiran: persentase, donut chart, ranking kelas."""
    from datetime import date, timedelta

    periode = request.args.get('periode', 'bulan')

    # Tentukan rentang tanggal berdasarkan periode
    today = date.today()
    if periode == 'bulan':
        tanggal_dari  = today.replace(day=1).isoformat()
        tanggal_sampai = today.isoformat()
    elif periode == 'semester':
        # Semester berjalan: Januari–Juni atau Juli–Desember
        bulan = today.month
        if bulan <= 6:
            tanggal_dari = today.replace(month=1, day=1).isoformat()
        else:
            tanggal_dari = today.replace(month=7, day=1).isoformat()
        tanggal_sampai = today.isoformat()
    else:  # tahun
        tanggal_dari  = today.replace(month=1, day=1).isoformat()
        tanggal_sampai = today.isoformat()

    # Data laporan
    persen        = db.get_persentase_kehadiran(tanggal_dari=tanggal_dari, tanggal_sampai=tanggal_sampai)
    ranking_kelas = db.get_ranking_kelas(tanggal_dari, tanggal_sampai)
    top_mahasiswa = db.get_top_mahasiswa(tanggal_dari, tanggal_sampai)
    statistik     = db.get_statistik_dashboard()

    return render_template('laporan/index.html',
                           active_page='laporan',
                           periode=periode,
                           persen=persen,
                           ranking_kelas=ranking_kelas,
                           top_mahasiswa=top_mahasiswa,
                           total_kelas=statistik.get('total_kelas', 0),
                           total_mahasiswa=statistik.get('total_mahasiswa', 0))


# ══════════════════════════════════════════════════════════════
# TAHAP 10: ABSENSI MANUAL ADMIN
# ══════════════════════════════════════════════════════════════

@app.route('/absensi/manual', methods=['GET', 'POST'])
@login_required
def absensi_manual():
    """Input / override absensi oleh admin."""
    if request.method == 'POST':
        absensi_id = request.form.get('absensi_id', type=int)
        status_baru = request.form.get('status')

        if not absensi_id or status_baru not in ('hadir', 'terlambat', 'izin', 'alpha'):
            flash('Data tidak valid.', 'error')
            return redirect(url_for('absensi_manual'))

        if db.update_status_absensi(absensi_id, status_baru):
            flash(f'Status absensi berhasil diubah ke "{status_baru}".', 'success')
        else:
            flash('Gagal mengubah status absensi.', 'error')
        return redirect(url_for('absensi_manual'))

    # GET — tampilkan daftar absensi hari ini untuk dioverride
    rekap = db.get_absensi_hari_ini()
    return render_template('absensi/manual.html',
                           active_page='rekap',
                           rekap=rekap)


# ══════════════════════════════════════════════════════════════
# API ENDPOINTS
# ══════════════════════════════════════════════════════════════

@app.route('/api/absensi/hari-ini')
@login_required
def api_absensi_hari_ini():
    """Data absensi hari ini dalam format JSON."""
    try:
        data = db.get_absensi_hari_ini()
        # Konversi timedelta/time ke string agar JSON-serializable
        for row in data:
            for key, val in row.items():
                if isinstance(val, timedelta):
                    total_seconds = int(val.total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    row[key] = f'{hours:02d}:{minutes:02d}:{seconds:02d}'
                elif isinstance(val, (date,)):
                    row[key] = val.isoformat()
        return jsonify({'status': 'ok', 'data': data, 'pesan': None})
    except Exception as e:
        return jsonify({'status': 'error', 'data': [], 'pesan': str(e)})


@app.route('/api/absensi/manual', methods=['POST'])
@login_required
def api_absensi_manual():
    """Catat absensi manual oleh admin (izin, sakit, hadir, dll)."""
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'pesan': 'Data tidak valid.'}), 400

    user_id = data.get('user_id')
    jadwal_id = data.get('jadwal_id')
    status = data.get('status', '').strip()
    alasan = data.get('alasan', '').strip() or None

    if not user_id or not jadwal_id or not status:
        return jsonify({'status': 'error', 'pesan': 'User, jadwal, dan status wajib diisi.'}), 400

    if status not in ('hadir', 'terlambat', 'izin', 'sakit', 'alpha'):
        return jsonify({'status': 'error', 'pesan': 'Status tidak valid.'}), 400

    # Validasi user dan jadwal ada
    user = db.get_user_by_id(int(user_id))
    if not user:
        return jsonify({'status': 'error', 'pesan': 'Mahasiswa tidak ditemukan.'}), 404

    tanggal = date.today()
    hasil = db.catat_absensi_manual(int(user_id), int(jadwal_id), tanggal, status, alasan)

    if hasil:
        aksi = 'diperbarui' if hasil['aksi'] == 'update' else 'dicatat'
        pesan = f"Absensi {user['nama']} berhasil {aksi} sebagai {status}."
        if hasil.get('status_lama'):
            pesan += f" (sebelumnya: {hasil['status_lama']})"
        print(f"[MANUAL] {pesan}")
        return jsonify({'status': 'ok', 'pesan': pesan, 'data': hasil})
    else:
        return jsonify({'status': 'error', 'pesan': 'Gagal mencatat absensi.'}), 500


@app.route('/api/mahasiswa/list')
@login_required
def api_mahasiswa_list():
    """Daftar semua mahasiswa untuk dropdown."""
    users = db.get_semua_user()
    data = [{'id': u['id'], 'nama': u['nama'], 'nim': u['nim'],
             'kelas_id': u['kelas_id'], 'nama_kelas': u['nama_kelas']} for u in users]
    return jsonify({'status': 'ok', 'data': data})


@app.route('/api/jadwal/hari-ini')
@login_required
def api_jadwal_hari_ini():
    """Daftar jadwal hari ini (semua, bukan hanya yang aktif)."""
    hari = _get_nama_hari()
    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT j.id, j.hari, j.jam_mulai, j.jam_selesai,
                   m.nama_mk, m.kelas_id, k.nama_kelas
            FROM jadwal j
            JOIN matakuliah m ON j.matakuliah_id = m.id
            JOIN kelas k ON m.kelas_id = k.id
            WHERE j.hari = %s
            ORDER BY j.jam_mulai
        """, (hari,))
        hasil = cursor.fetchall()
        # Konversi timedelta ke string
        for row in hasil:
            for key, val in row.items():
                if isinstance(val, timedelta):
                    total = int(val.total_seconds())
                    h, r = divmod(total, 3600)
                    m, s = divmod(r, 60)
                    row[key] = f'{h:02d}:{m:02d}'
        cursor.close(); conn.close()
        return jsonify({'status': 'ok', 'data': hasil})
    except Exception as e:
        return jsonify({'status': 'error', 'data': [], 'pesan': str(e)})

@app.route('/api/foto/upload', methods=['POST'])
@login_required
def api_foto_upload():
    """Terima foto wajah via AJAX (base64) dan simpan ke dataset/."""
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'pesan': 'Data tidak valid.'}), 400

    nama     = data.get('nama', '').strip()
    nim      = data.get('nim', '').strip()
    kelas_id = data.get('kelas_id')
    foto_b64 = data.get('foto', '')
    index    = data.get('index', 0)

    if not nama or not nim or not kelas_id or not foto_b64:
        return jsonify({'status': 'error', 'pesan': 'Field tidak lengkap.'}), 400

    try:
        # Pastikan user sudah ada di DB, buat jika belum
        user = db.get_user_by_nim(nim)
        if not user:
            user_id = db.tambah_user(nama, nim, int(kelas_id))
            if not user_id:
                return jsonify({'status': 'error', 'pesan': 'NIM sudah terdaftar atau gagal simpan.'}), 400
        else:
            user_id = user['id']

        # Decode base64 → numpy array
        header, encoded = foto_b64.split(',', 1)
        foto_bytes = base64.b64decode(encoded)
        np_arr = np.frombuffer(foto_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        if frame is not None:
            h, w = frame.shape[:2]
            # Crop area tengah (sesuai panduan oval: ~37.5% width, ~62.5% height)
            crop_w, crop_h = int(w * 0.45), int(h * 0.70)
            x1 = (w - crop_w) // 2
            y1 = (h - crop_h) // 2
            cropped = frame[y1:y1+crop_h, x1:x1+crop_w]
            # Simpan hasil crop
            save_frame = cropped
        else:
            # Fallback: simpan raw bytes
            save_frame = None

        folder = os.path.join(DATASET_PATH, str(user_id))
        os.makedirs(folder, exist_ok=True)

        # Cek batas maksimum foto — tolak jika sudah cukup
        from config import FOTO_PER_USER
        foto_tersimpan = len([f for f in os.listdir(folder) if f.endswith('.jpg')])
        if foto_tersimpan >= FOTO_PER_USER:
            return jsonify({
                'status': 'selesai',
                'pesan': f'Batas {FOTO_PER_USER} foto sudah tercapai. Silakan mulai training.',
                'data': {'user_id': user_id, 'total': foto_tersimpan}
            })

        filepath = os.path.join(folder, f'{index}.jpg')

        if save_frame is not None:
            cv2.imwrite(filepath, save_frame)
        else:
            with open(filepath, 'wb') as f:
                f.write(foto_bytes)

        return jsonify({'status': 'ok', 'pesan': f'Foto {index} tersimpan.', 'data': {'user_id': user_id}})

    except Exception as e:
        return jsonify({'status': 'error', 'pesan': str(e)}), 500


@app.route('/api/training/start', methods=['POST'])
@login_required
def api_training_start():
    """Mulai training LBPH di background thread."""

    def run_training():
        """Jalankan training di background."""
        try:
            from face.trainer import train_model
            train_model()
        except Exception as e:
            print(f'[ERROR] Training gagal: {e}')

    thread = threading.Thread(target=run_training, daemon=True)
    thread.start()

    return jsonify({
        'status': 'ok',
        'pesan': 'Training dimulai di background. Model akan diperbarui otomatis.',
        'data': None
    })


@app.route('/api/search')
@login_required
def api_search():
    """Endpoint API untuk mencari data mahasiswa dan jadwal secara realtime."""
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({
            'status': 'ok',
            'data': {'mahasiswa': [], 'jadwal': []},
            'pesan': 'Query pencarian kosong.'
        })

    try:
        mahasiswa = db.cari_mahasiswa(query)
        jadwal = db.cari_jadwal(query)

        # Konversi tipe data non-JSON-serializable (seperti timedelta) ke string
        for row in jadwal:
            for key in ['jam_mulai', 'jam_selesai', 'batas_terlambat']:
                if key in row:
                    val = row[key]
                    if isinstance(val, timedelta):
                        total_seconds = int(val.total_seconds())
                        hours, remainder = divmod(total_seconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        row[key] = f'{hours:02d}:{minutes:02d}'
                    elif hasattr(val, 'strftime'):
                        row[key] = val.strftime('%H:%M')
                    else:
                        row[key] = str(val)[:5]

        return jsonify({
            'status': 'ok',
            'data': {
                'mahasiswa': mahasiswa,
                'jadwal': jadwal
            },
            'pesan': 'Pencarian berhasil.'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'data': {'mahasiswa': [], 'jadwal': []},
            'pesan': f'Terjadi kesalahan saat mencari: {str(e)}'
        })




# ══════════════════════════════════════════════════════════════
# FACE RECOGNITION + ABSENSI OTOMATIS
# ══════════════════════════════════════════════════════════════

def _decode_frame(frame_b64):
    """Decode base64 frame menjadi numpy array BGR OpenCV."""
    try:
        # Hapus header data URI jika ada
        if ',' in frame_b64:
            frame_b64 = frame_b64.split(',', 1)[1]
        img_bytes = base64.b64decode(frame_b64)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        return frame
    except Exception as e:
        print(f'[ERROR] Gagal decode frame: {e}')
        return None


def _simpan_snapshot(frame, user_id):
    """Simpan snapshot bukti absensi ke folder snapshots/."""
    try:
        os.makedirs(SNAPSHOT_PATH, exist_ok=True)
        now = now_wib()
        filename = f"{user_id}_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
        filepath = os.path.join(SNAPSHOT_PATH, filename)
        cv2.imwrite(filepath, frame)
        return filepath
    except Exception as e:
        print(f'[ERROR] Gagal simpan snapshot: {e}')
        return None


def _kirim_ke_esp32(nama, nim, status_pesan):
    """Kirim notifikasi ke ESP32 via HTTP POST (jika diaktifkan)."""
    if not ESP32_ENABLED:
        return
    try:
        import requests
        url = f"http://{ESP32_IP}:{ESP32_PORT}/absensi"
        payload = {'nama': nama, 'nim': nim, 'status': status_pesan}
        requests.post(url, json=payload, timeout=ESP32_TIMEOUT)
        print(f'[ESP32] Notifikasi terkirim: {nama} - {status_pesan}')
    except Exception as e:
        print(f'[ESP32] Gagal kirim: {e}')


def _get_nama_hari():
    """Dapatkan nama hari ini dalam Bahasa Indonesia (timezone WIB)."""
    hari_map = {
        0: 'Senin', 1: 'Selasa', 2: 'Rabu',
        3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'
    }
    return hari_map.get(now_wib().weekday(), '')


def _proses_recognition(frame):
    """Proses satu frame: anti-spoofing → recognition → absensi.

    Alur lengkap sesuai context.md bagian 5.4:
    1. Cek anti-spoofing
    2. Predict wajah dengan LBPH
    3. Cari jadwal aktif hari ini
    4. Cek duplikasi absensi
    5. Tentukan status (hadir/terlambat)
    6. Simpan snapshot + catat absensi
    7. Kirim ke ESP32

    Returns:
        dict hasil proses untuk dikirim ke client
    """
    from face.anti_spoofing import check as spoofing_check
    from face.recognition import predict_single

    # ── 1. Anti-spoofing ──
    spoof_result = spoofing_check(frame)

    if not spoof_result['is_real']:
        # Spoofing terdeteksi — simpan bukti ke spoofing_log
        snapshot = _simpan_snapshot(frame, 'spoofing')
        db.catat_spoofing(snapshot, spoof_result['score'])
        return {
            'status': 'error',
            'tipe': 'spoofing',
            'score': spoof_result['score'],
            'pesan': 'Spoofing terdeteksi! Gunakan wajah asli.',
            'spoofing': spoof_result
        }

    if spoof_result['label'] == 'NO_FACE':
        return {
            'status': 'skip',
            'tipe': 'no_face',
            'pesan': 'Tidak ada wajah terdeteksi.',
            'spoofing': spoof_result
        }

    # ── 2. Predict wajah dengan LBPH ──
    result = predict_single(frame)

    if result is None:
        return {
            'status': 'skip',
            'tipe': 'no_face',
            'pesan': 'Wajah tidak terdeteksi untuk recognition.',
            'spoofing': spoof_result
        }

    # Log confidence untuk debug
    print(f'[ABSENSI] Predict: user_id={result["user_id"]}, '
          f'conf={result["confidence"]:.1f}, dikenali={result["dikenali"]}')

    if not result['dikenali']:
        # Reset counter konsekutif saat wajah tidak dikenali
        _consecutive_tracker['user_id'] = None
        _consecutive_tracker['count'] = 0
        return {
            'status': 'error',
            'tipe': 'unknown',
            'confidence': result['confidence'],
            'pesan': f'Wajah tidak dikenali (confidence: {result["confidence"]}).',
            'spoofing': spoof_result
        }

    # ── Verifikasi konsekutif: harus 3x berturut-turut user yang SAMA ──
    detected_uid = result['user_id']
    if _consecutive_tracker['user_id'] == detected_uid:
        _consecutive_tracker['count'] += 1
    else:
        _consecutive_tracker['user_id'] = detected_uid
        _consecutive_tracker['count'] = 1

    required = 3  # Minimal 3 frame konsekutif
    if _consecutive_tracker['count'] < required:
        return {
            'status': 'skip',
            'tipe': 'verifying',
            'pesan': f'Memverifikasi wajah... ({_consecutive_tracker["count"]}/{required})',
            'spoofing': spoof_result
        }

    # Reset counter setelah berhasil verifikasi
    _consecutive_tracker['user_id'] = None
    _consecutive_tracker['count'] = 0

    user_id = result['user_id']
    confidence = result['confidence']

    # ── 3. Ambil data mahasiswa ──
    user = db.get_user_by_id(user_id)
    if not user:
        return {
            'status': 'error',
            'tipe': 'user_not_found',
            'pesan': f'User ID {user_id} tidak ditemukan di database.',
            'spoofing': spoof_result
        }

    # ── 4. Cari jadwal aktif hari ini ──
    hari = _get_nama_hari()
    waktu_sekarang = now_wib().strftime('%H:%M:%S')
    jadwal_list = db.get_jadwal_aktif(hari, waktu_sekarang)

    if not jadwal_list:
        return {
            'status': 'error',
            'tipe': 'no_jadwal',
            'pesan': f'Tidak ada jadwal aktif saat ini ({hari} {waktu_sekarang}).',
            'data': {'nama': user['nama'], 'nim': user['nim']},
            'spoofing': spoof_result
        }

    # Cari jadwal yang sesuai kelas mahasiswa
    jadwal = None
    for j in jadwal_list:
        if j['kelas_id'] == user['kelas_id']:
            jadwal = j
            break

    if not jadwal:
        return {
            'status': 'error',
            'tipe': 'no_jadwal',
            'pesan': f'Tidak ada jadwal untuk kelas {user.get("nama_kelas", "")} saat ini.',
            'data': {'nama': user['nama'], 'nim': user['nim']},
            'spoofing': spoof_result
        }

    # ── 5. Cek duplikasi absensi ──
    tanggal_hari_ini = date.today()
    sudah = db.cek_sudah_absen(user_id, jadwal['id'], tanggal_hari_ini)

    if sudah:
        # Kirim notifikasi duplikat ke ESP32
        _kirim_ke_esp32(user['nama'], user['nim'], 'duplikat')
        return {
            'status': 'error',
            'tipe': 'duplikat',
            'pesan': f'{user["nama"]} sudah absen untuk {jadwal["nama_mk"]} hari ini.',
            'data': {
                'nama': user['nama'], 'nim': user['nim'],
                'status_absensi': sudah['status']
            },
            'spoofing': spoof_result
        }

    # ── 6. Tentukan status: hadir atau terlambat ──
    batas_str = str(jadwal['batas_terlambat'])
    # Konversi timedelta ke string waktu jika perlu
    if isinstance(jadwal['batas_terlambat'], timedelta):
        total_sec = int(jadwal['batas_terlambat'].total_seconds())
        h, m, s = total_sec // 3600, (total_sec % 3600) // 60, total_sec % 60
        batas_str = f'{h:02d}:{m:02d}:{s:02d}'

    status_absensi = 'hadir' if waktu_sekarang <= batas_str else 'terlambat'

    # ── 7. Simpan snapshot bukti absensi ──
    snapshot_path = _simpan_snapshot(frame, user_id)

    # ── 8. Catat absensi ke database ──
    absensi_id = db.catat_absensi(
        user_id=user_id,
        jadwal_id=jadwal['id'],
        tanggal=tanggal_hari_ini,
        waktu_absen=waktu_sekarang,
        status=status_absensi,
        snapshot_path=snapshot_path,
        dibuat_manual=False
    )

    if not absensi_id:
        return {
            'status': 'error',
            'tipe': 'db_error',
            'pesan': 'Gagal menyimpan absensi ke database.',
            'spoofing': spoof_result
        }

    # ── 9. Kirim ke ESP32 ──
    esp_status = 'berhasil'
    _kirim_ke_esp32(user['nama'], user['nim'], esp_status)

    # ── 10. Siapkan response ──
    data_response = {
        'nama': user['nama'],
        'nim': user['nim'],
        'nama_kelas': user.get('nama_kelas', ''),
        'nama_mk': jadwal['nama_mk'],
        'confidence': confidence,
        'status_absensi': status_absensi,
        'waktu_absen': waktu_sekarang,
        'absensi_id': absensi_id,
        'status': status_absensi
    }

    # Ambil statistik terbaru untuk update dashboard
    stats = db.get_statistik_dashboard()

    return {
        'status': 'ok',
        'pesan': f'Absensi {user["nama"]} berhasil ({status_absensi}).',
        'data': data_response,
        'stats': {
            'hadir': stats.get('hadir_hari_ini', 0),
            'terlambat': stats.get('terlambat_hari_ini', 0),
            'alpha': stats.get('alpha_hari_ini', 0)
        },
        'spoofing': spoof_result
    }


@app.route('/api/absensi/proses', methods=['POST'])
@login_required
def api_absensi_proses():
    """Proses frame dari kamera untuk face recognition + absensi.

    Menerima base64 frame, jalankan anti-spoofing → recognition → catat absensi.
    """
    data = request.get_json()
    if not data or 'frame' not in data:
        return jsonify({'status': 'error', 'pesan': 'Frame tidak ditemukan.', 'data': None}), 400

    frame = _decode_frame(data['frame'])
    if frame is None:
        return jsonify({'status': 'error', 'pesan': 'Gagal decode frame.', 'data': None}), 400

    hasil = _proses_recognition(frame)
    return jsonify(hasil)


@app.route('/api/camera/toggle', methods=['POST'])
@login_required
def api_camera_toggle():
    """Toggle status kamera ON/OFF."""
    data = request.get_json()
    if data and 'active' in data:
        camera_state['active'] = bool(data['active'])
    else:
        camera_state['active'] = not camera_state['active']

    status = 'on' if camera_state['active'] else 'off'
    return jsonify({
        'status': 'ok',
        'pesan': f'Kamera {status}.',
        'data': {'camera_active': camera_state['active']}
    })


# ══════════════════════════════════════════════════════════════
# SERVE SNAPSHOT — Menyajikan foto bukti absensi
# ══════════════════════════════════════════════════════════════

@app.route('/snapshots/<path:filename>')
@login_required
def serve_snapshot(filename):
    """Sajikan file gambar snapshot bukti absensi."""
    from flask import send_from_directory
    return send_from_directory(SNAPSHOT_PATH, filename)


# ══════════════════════════════════════════════════════════════
# WEBSOCKET HANDLERS (Flask-SocketIO)
# ══════════════════════════════════════════════════════════════

@socketio.on('connect')
def handle_connect():
    """Client terhubung via WebSocket."""
    print('[SOCKET] Client terhubung.')


@socketio.on('disconnect')
def handle_disconnect():
    """Client terputus."""
    print('[SOCKET] Client terputus.')


@socketio.on('camera_toggle')
def handle_camera_toggle(data):
    """Toggle kamera dari client."""
    camera_state['active'] = data.get('active', False)
    emit('camera_status', {'active': camera_state['active']}, broadcast=True)


@socketio.on('process_frame')
def handle_process_frame(data):
    """Terima frame dari client via WebSocket, proses recognition."""
    try:
        if 'frame' not in data:
            emit('recognition_result', {'status': 'error', 'pesan': 'Frame kosong.'})
            return

        frame = _decode_frame(data['frame'])
        if frame is None:
            emit('recognition_result', {'status': 'error', 'pesan': 'Gagal decode.'})
            return

        print(f'[SOCKET] process_frame: frame shape={frame.shape}')
        hasil = _proses_recognition(frame)
        print(f'[SOCKET] process_frame result: status={hasil.get("status")}, tipe={hasil.get("tipe", "-")}')
        emit('recognition_result', hasil)

        # Jika absensi berhasil, broadcast ke semua client
        if hasil.get('status') == 'ok' and hasil.get('data'):
            socketio.emit('absensi_update', {
                **hasil['data'],
                'stats': hasil.get('stats')
            })
    except Exception as e:
        print(f'[SOCKET ERROR] process_frame exception: {e}')
        import traceback
        traceback.print_exc()
        emit('recognition_result', {'status': 'error', 'pesan': f'Server error: {str(e)}'})



# ══════════════════════════════════════════════════════════════
# AUTO-ALPHA: Background thread untuk tandai alpha otomatis
# ══════════════════════════════════════════════════════════════

def _auto_alpha_checker():
    """Background thread: cek jadwal yang sudah selesai, tandai alpha otomatis."""
    import time as _time
    print('[AUTO-ALPHA] Background checker dimulai.')

    while True:
        _time.sleep(60)  # Cek setiap 60 detik
        try:
            hari = _get_nama_hari()
            waktu_sekarang = now_wib().strftime('%H:%M:%S')
            tanggal = now_wib().date()

            # Ambil semua jadwal yang sudah selesai hari ini
            jadwal_selesai = db.get_jadwal_selesai_hari_ini(hari, waktu_sekarang)

            for j in jadwal_selesai:
                jadwal_id = j['id']
                kelas_id = j['kelas_id']

                # Cari mahasiswa yang belum absen
                belum_absen = db.get_mahasiswa_belum_absen(jadwal_id, kelas_id, tanggal)

                if belum_absen:
                    user_ids = [m['id'] for m in belum_absen]
                    count = db.bulk_catat_alpha(jadwal_id, user_ids, tanggal)
                    if count > 0:
                        nama_mk = j.get('nama_mk', '?')
                        print(f'[AUTO-ALPHA] {count} mahasiswa ditandai alpha '
                              f'untuk {nama_mk} (jadwal #{jadwal_id})')

        except Exception as e:
            print(f'[AUTO-ALPHA] Error: {e}')


# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════


if __name__ == '__main__':
    # Pastikan folder yang dibutuhkan ada
    os.makedirs(SNAPSHOT_PATH, exist_ok=True)
    os.makedirs(DATASET_PATH, exist_ok=True)
    os.makedirs(os.path.dirname(MODEL_PATH), exist_ok=True)

    # Jalankan auto-alpha checker di background thread
    import threading
    alpha_thread = threading.Thread(target=_auto_alpha_checker, daemon=True)
    alpha_thread.start()

    print("=" * 50)
    print("   FLASK + SOCKETIO — SISTEM ABSENSI")
    print("=" * 50)
    print(f"\n[INFO] Dashboard  : http://127.0.0.1:{FLASK_PORT}")
    print(f"[INFO] Login      : http://127.0.0.1:{FLASK_PORT}/login")
    print(f"[INFO] WebSocket  : ws://127.0.0.1:{FLASK_PORT}")
    print(f"[INFO] Anti-Spoof : threshold={ANTI_SPOOFING_THRESHOLD}")
    print(f"[INFO] Confidence : threshold={CONFIDENCE_THRESHOLD}")
    print(f"[INFO] ESP32      : {'Aktif' if ESP32_ENABLED else 'Nonaktif'}")
    print(f"[INFO] Auto-Alpha : Aktif (cek setiap 60 detik)")
    print(f"[INFO] Tekan Ctrl+C untuk menghentikan.\n")
    socketio.run(app, host=FLASK_HOST, port=FLASK_PORT, debug=True,
                 allow_unsafe_werkzeug=True)